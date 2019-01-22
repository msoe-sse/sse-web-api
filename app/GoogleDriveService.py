import os
import json
import io
import openpyxl
from googleapiclient.discovery import build
from google.oauth2 import service_account
from googleapiclient.http import MediaIoBaseDownload
from app.GoogleServiceBuilder import GoogleServiceBuilder

SCOPES = ['https://www.googleapis.com/auth/drive']
FILE_ID = '1DXF7Qf6Zjc6yQIboDmqwl3s9l3Y6yKNK'

class GoogleDriveService():
    def __init__(self):
        self.drive_service = self._build_drive_service()
    
    def get_point_data(self):
        result = {}
        workbook = self._load_workbook()
        result['meetings'] = self._parse_meetings(workbook.active)
        result['students'] = self._parse_students(workbook.active)
        return result

    def _load_workbook(self):
        excel_bytes = self._download_excel_file()
        return  openpyxl.load_workbook(excel_bytes, data_only=True)
    
    def _download_excel_file(self):
        request = self.drive_service.files().get_media(fileId=FILE_ID)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
        return fh

    def _parse_meetings(self, worksheet):
        result = []
        for row in worksheet.iter_rows(max_row=1):
            for cell in row:
                if cell.value and cell.value != "TOTALS":
                    result.append(cell.value)
        return result
    
    def _parse_students(self, worksheet):
        result = []
        for row in worksheet.iter_rows(min_row=2):
            result.append(self._parse_student(row))
        return sorted(result, key = lambda x: (x['pointTotal'], x['name']), reverse=True)

    def _parse_student(self, row):
        current_student = {}
        current_student['name'] = row[0].value
        current_student['pointTotal'] = row[-1].value
        current_student['pointBreakdown'] = self._parse_point_breakdown(row)
        return current_student

    def _parse_point_breakdown(self, row):
        point_breakdown = []
        for i in range(1, len(row) - 2):
            value_to_append = 0
            if(row[i].value): value_to_append = row[i].value
            point_breakdown.append(value_to_append)
        return point_breakdown
    
    def _build_drive_service(self):
        service_builder = GoogleServiceBuilder()
        return service_builder.build_service(SCOPES, 'drive', 'v3')
